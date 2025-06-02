# Inputs_Cur.py
# You need to login to the Bloomberg Terminal for the script to work!
# All data is given in USD

import blpapi
import openpyxl
import shutil
import os
import time
from datetime import datetime
import traceback


def setup_bloomberg_session(ticker_symbol):
    """Initialize Bloomberg API session with detailed logging."""
    options = blpapi.SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)

    print(f"[INFO] Attempting to connect to Bloomberg for {ticker_symbol}...")
    if not session.start():
        print("[WARNING] Failed to start Bloomberg session. Ensure Bloomberg Terminal is running.")
        return None
    if not session.openService("//blp/refdata"):
        print("[WARNING] Failed to open Bloomberg reference data service.")
        session.stop()
        return None
    print("[INFO] Bloomberg session started successfully.")
    return session

def fetch_bloomberg_data(session, ticker, fields, field_to_name_map, start_year=2014, end_year=2024, timeout=30):
    """Fetch historical data from Bloomberg with timeout and error handling, using USD override."""
    if not fields:
        print("[INFO] No fields to fetch in this batch.")
        return {}

    if len(fields) > 25:
        # This is a hard limit for a single HistoricalDataRequest with many fields.
        # Consider splitting into multiple requests if more than 25 unique fields are needed at once.
        # For now, this script batches fields before calling this, so this check is an internal safeguard.
        print(f"[WARNING] Attempting to fetch {len(fields)} fields, which might exceed typical Bloomberg limits per request if not batched properly by calling function.")


    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")
    security = f"{ticker}" # Assumes ticker is correctly formatted like "AAPL US Equity"
    request.getElement("securities").appendValue(security)
    for field in fields:
        request.getElement("fields").appendValue(field)
    request.set("periodicitySelection", "YEARLY")
    request.set("startDate", f"{start_year}0101")
    request.set("endDate", f"{end_year}1231")

    # USD currency override
    overrides = request.getElement("overrides")
    override = overrides.appendElement()
    override.setElement("fieldId", "EQY_FUND_CRNCY")
    override.setElement("value", "USD")

    print(f"[DEBUG] Sending request for {security} with fields: {fields} (currency override: USD)")
    session.sendRequest(request)

    data = {field: {} for field in fields} # Initialize data structure for this batch
    invalid_fields_in_batch = []
    start_request_time = time.time()

    # Event handling loop
    while True:
        event = session.nextEvent(500) # Check for events every 500ms

        if time.time() - start_request_time > timeout:
            print(f"[WARNING] Timeout waiting for response for {security} after {timeout}s.")
            break # Exit loop on timeout

        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"[DEBUG] Bloomberg event timeout for {security}. Continuing to wait...")
            continue

        if event.eventType() in [blpapi.Event.RESPONSE, blpapi.Event.PARTIAL_RESPONSE]:
            for msg in event:
                # print(f"[DEBUG] Raw Bloomberg response message: {msg}") # Can be very verbose

                if msg.hasElement("responseError"):
                    error = msg.getElement("responseError")
                    error_message = error.getElement("message").getValueAsString()
                    print(f"[ERROR] Bloomberg API error for {security}: {error_message}")
                    # If responseError is present, it's likely a request-level issue, might affect all fields.
                    return {field: {year: "N/A (Request Error)" for year in range(start_year, end_year + 1)} for field in fields}


                if not msg.hasElement("securityData"):
                    print(f"[WARNING] No securityData element in response for {security}.")
                    continue

                security_data = msg.getElement("securityData")

                if security_data.hasElement("fieldExceptions"):
                    field_exceptions = security_data.getElement("fieldExceptions")
                    for j in range(field_exceptions.numValues()):
                        field_error = field_exceptions.getValueAsElement(j) # Use getValueAsElement
                        invalid_field_id = field_error.getElement("fieldId").getValueAsString()
                        error_info = field_error.getElement("errorInfo").getElement("message").getValueAsString()
                        field_name_display = field_to_name_map.get(invalid_field_id, "Unknown Field")
                        print(f"[WARNING] Invalid Bloomberg field: '{invalid_field_id}' (mapped to '{field_name_display}') for {security}. Error: {error_info}")
                        if invalid_field_id not in invalid_fields_in_batch:
                            invalid_fields_in_batch.append(invalid_field_id)

                if not security_data.hasElement("fieldData"):
                    print(f"[WARNING] No fieldData element in securityData for {security}.")
                    continue

                field_data_array = security_data.getElement("fieldData")
                # print(f"[DEBUG] Number of fieldData entries for {security}: {field_data_array.numValues()}")

                for k in range(field_data_array.numValues()):
                    datum = field_data_array.getValueAsElement(k) # Use getValueAsElement
                    date_obj = datum.getElement("date").getValueAsDatetime()
                    year = date_obj.year
                    # print(f"[DEBUG] Processing data for year {year} for {security}: {datum}") # Verbose

                    for field_id in fields: # Iterate over fields requested in *this batch*
                        if field_id in invalid_fields_in_batch:
                            data[field_id][year] = "N/A (Invalid Field)"
                            continue
                        if datum.hasElement(field_id):
                            try:
                                value = datum.getElement(field_id).getValueAsFloat()
                                data[field_id][year] = value
                                # print(f"[DEBUG] Fetched {field_id} for {year}: {value}") # Verbose
                            except blpapi.exception.ElementErrorException:
                                try: # Attempt to get as string if float fails
                                    value_str = datum.getElement(field_id).getValueAsString()
                                    data[field_id][year] = value_str
                                    # print(f"[DEBUG] Field {field_id} for year {year} is not a float, stored as string: {value_str}") # Verbose
                                except Exception as e_str:
                                    print(f"[WARNING] Could not get value for field {field_id} for year {year}: {e_str}")
                                    data[field_id][year] = "N/A (Error)"
                        else:
                            # Ensure year entry exists even if field is missing for that year in this datum
                            if year not in data[field_id]:
                                data[field_id][year] = None # Or "N/A (Missing)"
                                # print(f"[DEBUG] No data for {field_id} in {year} in this specific datum entry.") # Verbose


        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg in event:
                if msg.messageType() == blpapi.Name("SessionTerminated"):
                    print("[ERROR] Bloomberg session terminated unexpectedly.")
                    # Potentially raise an exception or handle reconnection if this is critical
                    return None # Indicates session failure

        if event.eventType() == blpapi.Event.RESPONSE:
            print(f"[INFO] Received final response for batch for {security}.")
            break # Exit loop after final response

    # Final check for fields that might not have received any data
    for field_id in fields:
        for year_chk in range(start_year, end_year + 1):
            if year_chk not in data[field_id]:
                data[field_id][year_chk] = "N/A (No Data)"
                print(f"[DEBUG] Field {field_id} for {year_chk} for {security} had no data, marked N/A.")


    if not any(data[field] for field in data if data[field] and isinstance(data[field], dict)): # Check if any field got actual data
        print(f"[WARNING] No data received for any requested field for {ticker} in this batch.")

    if invalid_fields_in_batch:
        print(f"[INFO] Bloomberg fields skipped or marked N/A due to invalidity for {security}: {invalid_fields_in_batch}")

    # print(f"[DEBUG] Final fetched data for this batch: {data}") # Can be very large
    return data


def calculate_derived_metrics(data, start_year=2014, end_year=2024):
    """Calculate derived metrics like DSO."""
    derived = {
        "DSO": {}
    }

    def get_val(source_field_code, year, default=0.0): # Ensure float for calculations
        val = data.get(source_field_code, {}).get(year)
        if isinstance(val, (int, float)):
            return float(val)
        # Try to convert if string representation of number
        if isinstance(val, str):
            try:
                return float(val)
            except ValueError:
                pass # Fall through to default if conversion fails
        return default


    for year in range(start_year, end_year + 1):
        revenue = get_val("SALES_REV_TURN", year)
        ar = get_val("BS_ACCT_NOTE_RCV", year) # Accounts Receivable
        derived["DSO"][year] = (ar / revenue * 365) if revenue != 0 else 0.0 # Avoid division by zero

    return derived

# --- FIELD MAP (Keep as is) ---
field_map = {
    # Income Statement (IS)
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS"},
    "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS"},
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "ARDR_DEPRECIATION_AMORTIZATION", "statement": "IS"},
    "Depreciation Expense": {"source": "BDH", "field": "ARDR_DEPRECIATION_EXP", "statement": "IS"},
    "Amortization Expense": {"source": "BDH", "field": "ARDR_AMORT_EXP", "statement": "IS"},
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPER_INC", "statement": "IS"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS"},
    "EPS Basic": {"source": "BDH", "field": "BASIC_EPS", "statement": "IS"},
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS", "statement": "IS"},
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_AVG_NUM_SH_FOR_EPS", "statement": "IS"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_SH_FOR_DILUTED_EPS", "statement": "IS"},

    # Balance Sheet (BS)
    "Cash & Cash Equivalents": {"source": "BDH", "field": "BS_CASH_NEAR_CASH_ITEM", "statement": "BS"},
    "Short-Term Investments": {"source": "BDH", "field": "BS_MKT_SEC_OTHER_ST_INVEST", "statement": "BS"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCT_NOTE_RCV", "statement": "BS"},
    "Inventory": {"source": "BDH", "field": "BS_INVENTORIES", "statement": "BS"},
    "Current Assets": {"source": "BDH", "field": "BS_CUR_ASSET_REPORT", "statement": "BS"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_GROSS_FIX_ASSET", "statement": "BS"},
    "Accumulated Depreciation": {"source": "BDH", "field": "BS_ACCUM_DEPR", "statement": "BS"},
    "Intangibles": {"source": "BDH", "field": "BS_DISCLOSED_INTANGIBLES", "statement": "BS"},
    "Goodwill": {"source": "BDH", "field": "BS_GOODWILL", "statement": "BS"},
    "Non-Current Assets": {"source": "BDH", "field": "BS_TOT_NON_CUR_ASSET", "statement": "BS"},
    "Accounts Payable": {"source": "BDH", "field": "BS_ACCT_PAYABLE", "statement": "BS"},
    "Short-Term Borrowings": {"source": "BDH", "field": "SHORT_TERM_DEBT_DETAILED", "statement": "BS"},
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "ST_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS"},
    "Long-Term Borrowings": {"source": "BDH", "field": "LONG_TERM_BORROWINGS_DETAILED", "statement": "BS"},
    "Long-Term Operating Lease Liabilities": {"source": "BDH", "field": "LT_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Non-Current Liabilities": {"source": "BDH", "field": "NON_CUR_LIAB", "statement": "BS"},
    "Non-Controlling Interest": {"source": "BDH", "field": "MINORITY_NONCONTROLLING_INTEREST", "statement": "BS"},

    # Cash Flow Statement (CF)
    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_CHANGE_IN_INVENTORIES", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Pre-paid expeses and Other CA": {"source": "BDH", "field": "CF_CHANGE_IN_PREPAID_EXP", "statement": "CF", "section": "Operating"}, # Changed field
    "Increase (Decrease) in Accounts Payable": {"source": "BDH", "field": "CF_CHANGE_IN_ACCOUNTS_PAYABLE", "statement": "CF", "section": "Operating"},
    "Increase (Decrease) in Accrued Revenues and Other CL": {"source": "BDH", "field": "CF_CHANGE_IN_ACCRUED_LIABILITY", "statement": "CF", "section": "Operating"}, # Changed field
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMPENSATION", "statement": "CF", "section": "Operating"},
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating"},
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "ACQUIS_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "DISPOSAL_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Acquisitions": {"source": "BDH", "field": "CF_CASH_FOR_ACQUIS_SUBSIDIARIES", "statement": "CF", "section": "Investing"},
    "Divestitures": {"source": "BDH", "field": "CF_CASH_FOR_DIVESTITURES", "statement": "CF", "section": "Investing"},
    "Increase in LT Investment": {"source": "BDH", "field": "CF_INCR_INVEST", "statement": "CF", "section": "Investing"},
    "Decrease in LT Investment": {"source": "BDH", "field": "CF_DECR_INVEST", "statement": "CF", "section": "Investing"},
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing"},
    "Debt Borrowing": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PROCEEDS", "statement": "CF", "section": "Financing"},
    "Debt Repayment": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PAYMENT", "statement": "CF", "section": "Financing"},
    "Dividends": {"source": "BDH", "field": "CF_DVD_PAID", "statement": "CF", "section": "Financing"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "PROC_FR_REPURCH_EQTY_DETAILED", "statement": "CF", "section": "Financing"},
    "Financing Cash Flow": {"source": "BDH", "field": "CFF_ACTIVITIES_DETAILED", "statement": "CF", "section": "Financing"},
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_EFFECT_FOREIGN_EXCHANGES", "statement": "CF", "section": "All"},

    # Additional Fields (BS)
    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS"},
    "Total Debt": {"source": "BDH", "field": "SHORT_AND_LONG_TERM_DEBT", "statement": "BS"},
    "Preferred Stock": {"source": "BDH", "field": "PFD_EQTY_HYBRID_CAPITAL", "statement": "BS"},
    "Enterprise Value": {"source": "BDH", "field": "ENTERPRISE_VALUE", "statement": "BS"},
    "Total Borrowings": {"source": "BDH", "field": "TOT_BORROWINGS", "statement": "BS"},
    "Total Leases": {"source": "BDH", "field": "TOT_LEASE_LIAB", "statement": "BS"},
    "Net Debt": {"source": "BDH", "field": "NET_DEBT", "statement": "BS"},
    "Effective Tax Rate": {"source": "BDH", "field": "EFF_TAX_RATE", "statement": "BS"},

    # Derived Metrics
    "Changes in Net Working Capital": {"source": "derived", "field": "Changes in Net Working Capital", "statement": "BS"}, # Usually a CF item, but calculated from BS
    "DSO": {"source": "derived", "field": "DSO", "statement": "IS"}, # Calculated from IS (Revenue) and BS (AR)
    "DIH": {"source": "derived", "field": "DIH", "statement": "BS"}, # Days Inventory Held
    "DPO": {"source": "derived", "field": "DPO", "statement": "BS"}, # Days Payable Outstanding
    "Net Cash from Investments & Acquisitions": {"source": "derived", "field": "Net Cash from Investments & Acquisitions", "statement": "CF", "section": "Investing"},
    "Increase (Decrease) in Other": {"source": "derived", "field": "Increase (Decrease) in Other", "statement": "CF", "section": "Operating"},
}

# --- field_cell_map (Keep as is) ---
field_cell_map = {
    # Income Statement (IS)
    "Revenue (Sales)": "G6",
    "COGS (Cost of Goods Sold)": "G7",
    "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9",
    "R&D (Research & Development)": "G10",
    "Other Operating (Income) Expenses": "G11",
    "EBITDA": "G12",
    "D&A (Depreciation & Amortization)": "G13",
    "Depreciation Expense": "G14",
    "Amortization Expense": "G15",
    "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17",
    "Interest Expense": "G18",
    "Interest Income": "G19",
    "FX (Gain) Loss": "G20",
    "Other Non-Operating (Income) Expenses": "G21",
    "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23",
    "Net Income": "G24",
    "EPS Basic": "G25",
    "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27",
    "Diluted Weighted Average Shares": "G28",

    # Balance Sheet (BS)
    "Cash & Cash Equivalents": "G33",
    "Short-Term Investments": "G34",
    "Accounts Receivable": "G35",
    "Inventory": "G36",
    "Current Assets": "G38",
    "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41",
    "Intangibles": "G43",
    "Goodwill": "G44",
    "Non-Current Assets": "G47",
    "Accounts Payable": "G49",
    "Short-Term Borrowings": "G51",
    "Current Portion of Lease Liabilities": "G52",
    "Current Liabilities": "G54",
    "Long-Term Borrowings": "G56",
    "Long-Term Operating Lease Liabilities": "G57",
    "Non-Current Liabilities": "G59",
    "Non-Controlling Interest": "G62", # Already present, kept for consistency

    # Cash Flow Statement (CF)
    "(Increase) Decrease in Accounts Receivable": "G69",
    "(Increase) Decrease in Inventories": "G70",
    "(Increase) Decrease in Pre-paid expeses and Other CA": "G71",
    "Increase (Decrease) in Accounts Payable": "G72",
    "Increase (Decrease) in Accrued Revenues and Other CL": "G73",
    "Stock Based Compensation": "G74",
    "Operating Cash Flow": "G76",
    "Acquisition of Fixed & Intangibles": "G78",
    "Disposal of Fixed & Intangibles": "G79",
    "Acquisitions": "G81",
    "Divestitures": "G82",
    "Increase in LT Investment": "G83",
    "Decrease in LT Investment": "G84",
    "Investing Cash Flow": "G86",
    "Debt Borrowing": "G87", # Note: Changed from "G87" which was duplicated. This might need checking with template. Assuming G87 is Debt Borrowing.
    "Debt Repayment": "G88",
    "Dividends": "G90",
    "Increase (Repurchase) of Shares": "G91",
    "Financing Cash Flow": "G93",
    "Effect of Foreign Exchange": "G94",

    # Additional Fields (BS) - Note: some like Non-Controlling Interest are repeated, ensure template matches
    "Market Capitalization": "G99",
    "Total Debt": "G101",
    "Preferred Stock": "G102",
    # "Non-Controlling Interest": "G103", # This is also G62, check template if it's distinct or repeated
    "Enterprise Value": "G104",
    "Total Borrowings": "G100", # Assuming G100 based on typical order, was G96 previously
    "Total Leases": "G105", # Example, was G116
    "Net Debt": "G106", # Example, was G98
    "Effective Tax Rate": "G107", # Example, was G99

    # Derived Metrics (cells are usually where these are displayed/calculated in Excel, not direct inputs)
    "DSO": "G101", # This cell in your template is for the calculated DSO.
    # Other derived metrics usually have their own cells if displayed.
    # "Changes in Net Working Capital": "G100", # Often calculated within CF section or derived separately
    # "DIH": "G102",
    # "DPO": "G103",
    # "Net Cash from Investments & Acquisitions": "G82", # Could be sum of items in CF investing
    # "Increase (Decrease) in Other": "G72", # Could be a plug in CF operating
}


def filter_field_map_for_task(task_name, field_map_dict):
    """Filters the field_map for a specific task (e.g., 'IS', 'BS')."""
    statement_code, _ = task_name.split("_") if "_" in task_name else (task_name, None)

    allowed_statements = ["IS", "BS", "CF", "All"] # 'All' could be for meta/market data
    if statement_code not in allowed_statements:
        raise ValueError(f"Invalid statement code '{statement_code}'. Must be one of {allowed_statements}.")

    task_specific_fields = {}
    for name, config in field_map_dict.items():
        if config["statement"] == statement_code or config["statement"] == "All": # Include 'All' statement fields
            task_specific_fields[name] = config
    
    # Logic for including dependencies for derived fields (if any derived fields were part of the task)
    required_bdh_for_derived = set()
    for name, config in task_specific_fields.items():
        if config["source"] == "derived":
            # Example: if DSO needs Revenue and AR
            if name == "DSO":
                required_bdh_for_derived.update(["BS_ACCT_NOTE_RCV", "SALES_REV_TURN"])
            # Add more derived field dependencies here
    
    # Add these required BDH fields to the task if not already present
    for bdh_field_code in required_bdh_for_derived:
        found = any(c.get("field") == bdh_field_code and c.get("source") == "BDH" for c in task_specific_fields.values())
        if not found:
            # Find the global config for this BDH field and add it
            for global_name, global_config in field_map_dict.items():
                if global_config.get("field") == bdh_field_code and global_config.get("source") == "BDH":
                    # Add with a special prefix or flag if you need to distinguish these auto-added dependencies
                    task_specific_fields[f"__dep_{global_name}"] = global_config
                    break
    return task_specific_fields


def batch_fields(fields_to_fetch, batch_size=25):
    """Split fields into batches of batch_size or fewer."""
    unique_fields = sorted(list(set(fields_to_fetch))) # Ensure unique fields
    return [unique_fields[i:i + batch_size] for i in range(0, len(unique_fields), batch_size)]

def get_column_letter_from_index(col_index):
    """Convert 1-based column index to letter (e.g., 7 -> G)."""
    return openpyxl.utils.get_column_letter(col_index)

def get_target_cells_for_years(base_cell_ref, num_years):
    """Get list of cell references for a row, for num_years, starting from base_cell_ref."""
    try:
        col_str = "".join(filter(str.isalpha, base_cell_ref))
        row_num_str = "".join(filter(str.isdigit, base_cell_ref))
        if not col_str or not row_num_str: # Basic validation
            raise ValueError(f"Invalid cell reference format: {base_cell_ref}")
        row_num = int(row_num_str)
        start_col_idx = openpyxl.utils.column_index_from_string(col_str)

        target_cells = []
        for i in range(num_years):
            target_col_letter = get_column_letter_from_index(start_col_idx + i)
            target_cells.append(f"{target_col_letter}{row_num}")
        return target_cells
    except ValueError as e:
        print(f"[ERROR] Invalid base cell reference '{base_cell_ref}': {e}")
        raise # Re-raise to be handled by the caller or stop execution


def populate_valuation_model(template_path, output_directory, ticker_symbol_from_web):
    """Populate the 'Inputs' sheet with data for all statements in a single file."""
    if not os.path.exists(template_path):
        print(f"[ERROR] Template file '{template_path}' not found.")
        raise FileNotFoundError(f"Template file '{template_path}' not found.")

    # Sanitize ticker_symbol_from_web for use in filenames
    safe_ticker_name = ticker_symbol_from_web.replace(" ", "_").replace("/", "_").replace(":", "_")
    output_file_name = f"{safe_ticker_name}_Valuation_Model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    final_output_path = os.path.join(output_directory, output_file_name)

    os.makedirs(output_directory, exist_ok=True) # Ensure output directory exists

    shutil.copy(template_path, final_output_path)
    print(f"[INFO] Copied template '{template_path}' to output file '{final_output_path}'.")

    wb = openpyxl.load_workbook(final_output_path)
    if "Inputs" not in wb.sheetnames:
        print("[ERROR] 'Inputs' sheet not found in the workbook.")
        raise ValueError("'Inputs' sheet not found in the template file.")
    ws = wb["Inputs"]

    # Define years for data fetching and Excel population
    data_years = list(range(2014, 2025)) # For years 2014 through 2024 inclusive

    # Global dictionary to store all fetched BDH data to avoid redundant calls
    all_fetched_bdh_data = {}
    # Map Bloomberg field codes back to human-readable names for error reporting
    global_bberg_code_to_name_map = {
        config["field"]: name
        for name, config in field_map.items()
        if config["source"] == "BDH" and "field" in config
    }

    print(f"\n[PHASE] Starting data fetching for ticker: {ticker_symbol_from_web}")

    # Collect all unique BDH fields needed across all tasks
    all_required_bdh_codes = set()
    for task_name_iter in ["IS", "BS", "CF"]: # Iterate through main statements
        current_task_field_configs_iter = filter_field_map_for_task(task_name_iter, field_map)
        for name, config in current_task_field_configs_iter.items():
            if config["source"] == "BDH" and "field" in config:
                all_required_bdh_codes.add(config["field"])
    
    print(f"[INFO] Total unique Bloomberg fields to fetch: {len(all_required_bdh_codes)}")

    # Fetch all BDH data in batches
    field_batches = batch_fields(list(all_required_bdh_codes))

    session = None # Initialize session outside the loop
    try:
        session = setup_bloomberg_session(ticker_symbol_from_web)
        if not session:
            raise ConnectionError("Failed to establish Bloomberg session.")

        for batch_idx, current_batch_fields in enumerate(field_batches):
            print(f"  [BATCH {batch_idx + 1}/{len(field_batches)}] Fetching {len(current_batch_fields)} fields...")
            
            batch_data_fetched = fetch_bloomberg_data(
                session, ticker_symbol_from_web, current_batch_fields,
                global_bberg_code_to_name_map, start_year=min(data_years), end_year=max(data_years)
            )

            if batch_data_fetched: # Check if data was fetched
                for field_code, yearly_data in batch_data_fetched.items():
                    if field_code not in all_fetched_bdh_data:
                        all_fetched_bdh_data[field_code] = {}
                    # Merge yearly data, preferring new data if there's overlap (though should be unique)
                    all_fetched_bdh_data[field_code].update(yearly_data)
            else:
                print(f"    [WARNING] Batch {batch_idx + 1} returned no data.")
            
            # Optional: Add a small delay between batches if hitting rate limits, though session should handle some.
            # time.sleep(1) # Example: 1-second delay

    except ConnectionError as ce:
        print(f"[CRITICAL] Bloomberg Connection Error: {ce}")
        # Potentially re-raise or handle as a failure for the web app
        if session: session.stop()
        return None # Indicate failure
    except Exception as e:
        print(f"[CRITICAL] Error during data fetching: {e}")
        traceback.print_exc()
        if session: session.stop()
        return None # Indicate failure
    finally:
        if session:
            session.stop()
            print("[INFO] Bloomberg session stopped after fetching all batches.")


    print(f"\n[PHASE] Completed all data fetching.")
    # print(f"[DEBUG] All fetched BDH data keys: {list(all_fetched_bdh_data.keys())}")


    print(f"\n[PHASE] Calculating derived metrics...")
    all_derived_data = calculate_derived_metrics(all_fetched_bdh_data, start_year=min(data_years), end_year=max(data_years))
    print("[INFO] Derived metrics calculated.")

    print(f"\n[PHASE] Writing all data to Excel sheet '{ws.title}'...")

    for item_name, config in field_map.items():
        if item_name.startswith("__dep_"): # Skip auto-added dependency fields for direct Excel writing
            continue

        base_cell_ref = field_cell_map.get(item_name)
        if not base_cell_ref:
            # print(f"  [SKIP] No cell mapping for item: {item_name}") # Optional: for debugging missing mappings
            continue

        try:
            target_cells_for_item = get_target_cells_for_years(base_cell_ref, len(data_years))
        except ValueError as e:
            print(f"  [ERROR] Skipping item '{item_name}' due to cell reference error: {e}")
            continue


        for i, year_to_populate in enumerate(data_years):
            cell_ref_to_write = target_cells_for_item[i]
            value_to_write = 0.0 # Default to 0.0 for numbers

            if config["source"] == "BDH":
                bberg_field_code = config.get("field")
                if bberg_field_code:
                    data_source_for_item = all_fetched_bdh_data.get(bberg_field_code, {})
                    raw_value = data_source_for_item.get(year_to_populate)

                    if isinstance(raw_value, (int, float)):
                        value_to_write = float(raw_value)
                    elif isinstance(raw_value, str) and "N/A" in raw_value:
                        value_to_write = raw_value # Keep N/A as string
                    else: # Includes None or other unexpected types
                        value_to_write = 0.0 # Default for missing/unparseable BDH data
                        # print(f"  [DEBUG] BDH default for {item_name} ({bberg_field_code}) year {year_to_populate} = 0.0 (raw: {raw_value})")


            elif config["source"] == "derived":
                derived_field_name = config.get("field")
                if derived_field_name:
                    data_source_for_item = all_derived_data.get(derived_field_name, {})
                    value = data_source_for_item.get(year_to_populate)
                    if isinstance(value, (int, float)):
                        value_to_write = float(value)
                    else: # Includes None or other unexpected types
                        value_to_write = 0.0 # Default for missing/unparseable derived data
                        # print(f"  [DEBUG] Derived default for {item_name} ({derived_field_name}) year {year_to_populate} = 0.0 (raw: {value})")

            # Write to cell
            ws[cell_ref_to_write] = value_to_write
            # Apply number formatting if it's a number
            if isinstance(value_to_write, float):
                if "DSO" in item_name or "DIH" in item_name or "DPO" in item_name: # Specific format for day counts
                    ws[cell_ref_to_write].number_format = "0.0"
                else: # General number format
                    ws[cell_ref_to_write].number_format = '_(* #,##0.000_);_(* (#,##0.000);_(* "-"??_);_(@_)'


    wb.save(final_output_path)
    print(f"\n[SUCCESS] Valuation model populated and saved to '{final_output_path}'")
    return final_output_path


if __name__ == "__main__":
    print("-" * 70)
    print("Bloomberg Data to Excel Valuation Model Populator")
    print("-" * 70)
    print("This script fetches financial data for IS, BS, and CF statements,")
    print("processes it in batches, calculates derived metrics, and populates")
    print("a single Excel template.")
    print("Ensure Bloomberg Terminal is running and blpapi is correctly configured.")
    print("-" * 70)

    excel_template_path = "LIS_Valuation_Empty.xlsx"  # Make sure this file exists in the same directory or provide full path
    # output_directory = "."  # Save in the current directory by default
    # For web app, this will be passed by server.py, e.g., a 'static/outputs' folder
    default_web_output_dir = "currency_analyses_outputs_standalone_test"


    ticker_input_main = ""
    while not ticker_input_main:
        ticker_input_main = input("Enter the Ticker Symbol (e.g., AAPL US Equity for Apple Inc.): ").strip().upper()
        if not ticker_input_main:
            print("[VALIDATION] Ticker symbol cannot be empty. Please try again.")

    # Construct output path for standalone execution
    # output_file_name_main = f"{ticker_input_main.replace(' ', '_').replace('/', '_')}_Valuation_Model_{datetime.now().strftime('%Y%m%d')}.xlsx"
    # final_output_path_main = os.path.join(default_web_output_dir, output_file_name_main)
    
    print(f"\n[SETUP] Template: '{excel_template_path}'")
    # print(f"[SETUP] Output will be: '{final_output_path_main}'") # Path is now constructed inside the function
    print(f"[SETUP] Ticker: '{ticker_input_main}'")
    print(f"[SETUP] Default output directory for standalone test: '{default_web_output_dir}'")


    try:
        print("\nStarting the data population process (standalone test)...\n")
        # Call the main function with parameters
        generated_file = populate_valuation_model(
            template_path=excel_template_path,
            output_directory=default_web_output_dir, # Use a specific directory for testing
            ticker_symbol_from_web=ticker_input_main
        )
        if generated_file:
            print(f"\nProcess completed successfully. File saved to: {generated_file}")
        else:
            print("\nProcess failed to generate file.")

    except FileNotFoundError as e:
        print(f"[CRITICAL ERROR] The template Excel file was not found: {e}")
    except ConnectionError as e:
        print(f"[CRITICAL ERROR] Could not connect to Bloomberg: {e}")
        print("Please ensure the Bloomberg Terminal is running and API connectivity is configured.")
    except ValueError as e: # Catching specific ValueErrors like from cell ref issues
        print(f"[CRITICAL ERROR] A value-related error occurred: {e}")
    except Exception as e:
        print(f"[CRITICAL ERROR] An unexpected error occurred: {e}")
        # import traceback # Already imported at top
        print("\n--- Traceback ---")
        traceback.print_exc()
        print("--- End Traceback ---\n")
    finally:
        print("\nScript execution finished.")